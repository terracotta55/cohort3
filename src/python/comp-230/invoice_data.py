# from openpyxl import Workbook
from openpyxl import load_workbook
from invoice_functions import *

# load the excel file sheets
my_invoice_data = load_workbook('invoice_data.xlsx')
my_customers = my_invoice_data['customers']
my_invoices = my_invoice_data['invoices']
my_line_items = my_invoice_data['line_items']
my_product_list = my_invoice_data['product_list']

# create dictionary from sheets
customers_dict = sheet_to_dict(my_customers)
invoices_dict = sheet_to_dict(my_invoices)
line_items_dict = sheet_to_dict(my_line_items)
products_dict = sheet_to_dict(my_product_list)

invoice_id = input(
    'Enter invoice_id (choose from 101 - 110): ')
try:
    # find the id in invoices and create a new dictionary by joining all tables
    invoice_id = float(invoice_id)
    newInvoice = make_invoice(
        invoice_id, customers_dict, invoices_dict, line_items_dict, products_dict)
    # generating the invoice file
    invoiceFile = load_workbook('invoice_blank.xlsx')
    invoiceSheet = invoiceFile['invoice_sheet']
    invoiceSheet['F5'] = f"Invoice No.: {newInvoice['invoice_blank']['id']}"
    invoiceSheet['H3'] = newInvoice['invoice']['date']
    invoiceSheet['A7'] = f"{newInvoice['customer']['first_name']} {newInvoice['customer']['last_name']}\n{newInvoice['customer']['address']}"
    invoiceSheet['E7'] = f"{newInvoice['invoice']['shipping_address']}"
    row = 10
    for item in newInvoice['items']:
        invoiceSheet[f"A{row}"] = item['description']
        invoiceSheet[f"D{row}"] = item['quantity']
        invoiceSheet[f"E{row}"] = item['unit']
        invoiceSheet[f"F{row}"] = item['unit_price']
        invoiceSheet[f"H{row}"] = item['quantity'] * item['unit_price']
        row = row + 1
    invoiceFile.save('invoice_new.xlsx')
except:
    print('The invoice id you entered could not be found ...!')
