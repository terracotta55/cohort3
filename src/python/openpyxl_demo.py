from openpyxl import Workbook
from openpyxl import load_workbook
import json
'''
new_workbook = Workbook()
sheet_1 = new_workbook.active

sheet_1["A1"] = "write into"
sheet_1["B1"] = "excel worksheet"

new_workbook.save(filename="hello_world.xlsx")
'''
workbook = load_workbook(filename="openpyxl_demo.xlsx")
print(workbook.sheetnames)
sheet = workbook.active
# print(sheet)
print(sheet.title)
# print(sheet["A1"].value)
# print(sheet["F10"].value)
'''
for value in sheet.iter_rows(min_row=1, max_row=2, min_col=1, max_col=3, values_only=True):
    print(value)
'''
'''
for row in sheet.rows:
    print(row)
'''
for value in sheet.iter_rows(min_row=1,
                             max_row=1,
                             values_only=True):
    print(value)

for value in sheet.iter_rows(min_row=2,
                             min_col=4,
                             max_col=7,
                             values_only=True):
    print(value)

# put data into a dictionary
products = {}
# Using the values_only because you want to return the cells' values
for row in sheet.iter_rows(min_row=2,
                           min_col=4,
                           max_col=7,
                           values_only=True):
    product_id = row[0]
    product = {
        "parent": row[1],
        "title": row[2],
        "category": row[3]
    }
    products[product_id] = product

# Using json here to be able to format the output for displaying later
print(json.dumps(products))
